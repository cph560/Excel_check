import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import PatternFill
import logging
import logging.handlers


# 1.创建1个logger对象：

def init_log():
    
    
    log_path = os.getcwd() + "/log"
    try:
        if not os.path.exists(log_path):
            os.makedirs(log_path)
    except:
        print("创建日志目录失败")
        exit(1)
    if len(lg.handlers) == 0:  # 避免重复
        # 2.创建handler(负责输出，输出到屏幕streamhandler,输出到文件filehandler)
        filename = os.path.join(log_path, 'project.log')
        fh = logging.FileHandler(filename,mode="a",encoding="utf-8")#默认mode 为a模式，默认编码方式为utf-8
        sh = logging.StreamHandler()
        # 3.创建formatter：
        formatter=logging.Formatter(fmt='%(asctime)s - %(levelname)s - Model:%(filename)s - Fun:%(funcName)s - Message:%(message)s - Line:%(lineno)d')
        # 4.绑定关系：①logger绑定handler
        lg.addHandler(fh)
        lg.addHandler(sh)
        # ②为handler绑定formatter
        fh.setFormatter(formatter)
        sh.setFormatter(formatter)
        # 5.设置日志级别(日志级别两层关卡必须都通过，日志才能正常记录)
        lg.setLevel(40)
        fh.setLevel(40)
        sh.setLevel(40)

    return 

def check_update(File1_path, File2_path):

    df1 = pd.read_excel(File1_path)
    df2 = pd.read_excel(File2_path)

    if df1.shape[1] != df2.shape[1]:
        raise ValueError("比较文件中Column数量必须相同!!")
        
    titles = df1.keys()

    workEx_a = openpyxl.load_workbook(rf'{File1_path}')
    workEx_b = openpyxl.load_workbook(rf'{File2_path}')

    sheet_a = workEx_a['Sheet0']
    sheet_b = workEx_b['Sheet0']

    row = df1.shape[0]
    col = df1.shape[1]

    df1 = df1.fillna(value='NONE')
    df2 = df2.fillna(value='NONE')
    trigger = 0

    if df1.shape[0] == df2.shape[0]:
        comparsion = df1.values==df2.values

        for i in range(row):
            for j in range(col):
                if comparsion[i, j] == False:
                    trigger += 1
                    print(f'mismatched {df1.keys()[0]}({str(sheet_b.cell(i+2, 1).value)})')
                    
                    sheet_b.cell(i+2, 1).fill = PatternFill("solid", fgColor='FFFF00')
                    sheet_b.cell(i+2, j+1).fill = PatternFill("solid", fgColor='FFFF00')
                    sheet_b.cell(i+2, j+1).value = str(sheet_a.cell(i+2, j+1).value)+'-->'+str(sheet_b.cell(i+2, j+1).value)
        
        
        if trigger == 0:
            print('No Update on file')
        else:
            print('Checked File saved on Original Path')
            if os.path.exists('checked_File.xlsx'):
                os.remove(os.getcwd()+'\checked_File.xlsx')
            workEx_b.save(f'checked_File.xlsx')
    else:
        print(f"{File1_path} Reqs ----> {df1.shape[0]}, {File2_path} Reqs ----> {df2.shape[0]}"+'\n')

        dic1 = {}
        dic2 = {}
        for i in range(df1.shape[0]):
            data1 = df1.iloc[i, 0]
            dic1[data1] = []
            for j in range(1, df1.shape[1]):
                dic1[data1].append(df1.iloc[i, j])
            

        for i in range(df2.shape[0]):
            data2 = df2.iloc[i, 0]
            dic2[data2] = []
            for j in range(1, df2.shape[1]):
                dic2[data2].append(df2.iloc[i, j])
        
        res = []
        if len(dic2.keys())>len(dic1.keys()):
            but = 'Add'
            list_dic = list(dic2.keys())
        else:
            but = 'Delete'
            list_dic = list(dic1.keys())

        updated_req = 0
        for key in dic1.keys():
            if dic1[key] == dic2[key]:
                list_dic.remove(key)
                continue
            else:
                list_dic.remove(key)
                updated_req += 1
                for i in range(len(dic1[key])):
                    if dic1[key][i] != dic2[key][i]:
                        res.append(str(key)+"  ------>  "+str(titles[i+1])+ ": "+str(dic1[key][i])+"  ------>  "+str(dic2[key][i]))
             
        
        if len(res) == 0:
            print('No Update')
        else:
            print(f'Updated with total {updated_req} Reqs. from last version:')
            for val in res:
                print(f"{titles[0]}: "+val)
            print('\n')
            print(f'{but} total {len(list_dic)} new Requirments: ')
            for val2 in list_dic:
                print(f'{titles[0]}: {val2}')
    return 

if __name__== '__main__':
    # os.system('cmd /c "pip install --upgrade -r requirements.txt"')
    lg = logging.getLogger("Error")
    init_log()
    File1_path = input("Enter your xlsx file-Early Version:")
    File2_path = input("Enter your xlsx file-Last Version:")
    File1_path = rf'{File1_path}'
    File2_path = rf'{File2_path}'
    try:
        check_update(File1_path, File2_path)
    except Exception as e:
        lg.error(e)
    
    
    y=input("是否清空错误日志Y/N?\n").upper()
    if y=='Y':
        if(os.path.exists(os.getcwd() + "\log\project.log")):   # 判断生成的路径对不对，防止报错
            try:
                logging.shutdown()
                os.remove(os.getcwd() + "\log\project.log")
            except Exception as e:
                print(e)
        

    
    print("请问你是否需要退出？\n")
    i = 'N'
    while i!="Y": 
        i=input("请输入Y/N\n").upper()
#     if list1[i]!=list2[i]:
#         dict[i] =  
